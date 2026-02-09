import io
import re
from datetime import datetime, timedelta

import httpx
import openpyxl
import structlog
from tenacity import retry, stop_after_attempt, wait_exponential

from app.ingestion.base import BaseCollector

logger = structlog.get_logger()

# California EDD publishes WARN data as XLSX (updated Tue/Thu)
CA_WARN_URL = "https://edd.ca.gov/siteassets/files/jobs_and_training/warn/warn_report1.xlsx"
# Sheet name for detailed data (has trailing space in the actual file)
CA_SHEET_NAME = "Detailed WARN Report "

# Texas publishes WARN data via Socrata SODA API (no auth required)
TX_WARN_URL = "https://data.texas.gov/resource/8w53-c4f6.json"

# New York DOL WARN notices page
NY_WARN_URL = "https://dol.ny.gov/warn-notices"


class WarnActCollector(BaseCollector):
    source_name = "WARN Act (data.gov)"
    source_type = "warn_act"

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, max=16))
    async def _fetch_xlsx(self, url: str) -> bytes:
        async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
            response = await client.get(url)
            response.raise_for_status()
            return response.content

    async def collect(self) -> list[dict]:
        signals = []

        # California (existing)
        try:
            xlsx_bytes = await self._fetch_xlsx(CA_WARN_URL)
            ca_signals = self._parse_california_xlsx(xlsx_bytes)
            signals.extend(ca_signals)
            logger.info("warn_act.california_fetched", count=len(ca_signals))
        except Exception as e:
            logger.warning("warn_act.california_failed", error=str(e))

        # Texas (Socrata SODA API)
        try:
            tx_signals = await self._fetch_texas()
            signals.extend(tx_signals)
            logger.info("warn_act.texas_fetched", count=len(tx_signals))
        except Exception as e:
            logger.warning("warn_act.texas_failed", error=str(e))

        # New York (DOL HTML page)
        try:
            ny_signals = await self._fetch_new_york()
            signals.extend(ny_signals)
            logger.info("warn_act.new_york_fetched", count=len(ny_signals))
        except Exception as e:
            logger.warning("warn_act.new_york_failed", error=str(e))

        return signals

    def _parse_california_xlsx(self, xlsx_bytes: bytes) -> list[dict]:
        results = []
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True)

        # Find the detailed report sheet
        ws = None
        for name in wb.sheetnames:
            if "detailed" in name.lower() and "warn" in name.lower():
                ws = wb[name]
                break
        if ws is None:
            logger.warning("warn_act.sheet_not_found", sheets=wb.sheetnames)
            wb.close()
            return results

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            wb.close()
            return results

        # Row 0 is title, row 1 is headers
        # Headers: County/Parish, Notice Date, Processed Date, Effective Date, Company, Layoff/Closure, No. Of Employees, Address, Related Industry
        for row in rows[2:]:
            try:
                if len(row) < 7:
                    continue

                county = row[0]
                notice_date = row[1]
                effective_date = row[3]
                company_name = row[4]
                layoff_type = row[5]
                employees = row[6]
                address = row[7] if len(row) > 7 else ""

                if not company_name:
                    continue
                company_name = str(company_name).strip()

                # Parse employees
                emp_count = 0
                if employees is not None:
                    try:
                        emp_count = int(str(employees).replace(",", "").strip())
                    except (ValueError, TypeError):
                        emp_count = 0

                # Parse date
                event_date = None
                date_val = effective_date or notice_date
                if isinstance(date_val, datetime):
                    event_date = date_val.date()
                elif isinstance(date_val, str) and date_val.strip():
                    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
                        try:
                            event_date = datetime.strptime(date_val.strip(), fmt).date()
                            break
                        except ValueError:
                            continue

                # Extract city from address (format: "30825 Wiegman Road  Hayward CA 94544")
                # Use double-space as delimiter — EDD formats as "Street  City CA Zip"
                city = ""
                if address and isinstance(address, str):
                    # Split on double-space which separates street from city
                    addr_parts = re.split(r"\s{2,}", address.strip())
                    if len(addr_parts) >= 2:
                        # Last segment before zip is "City CA 94544"
                        city_state = addr_parts[-1].strip()
                        # Remove "CA" and zip
                        city = re.sub(r"\s+CA\s+\d{5}(-\d{4})?$", "", city_state).strip()
                    if not city:
                        # Fallback: take 1-2 words before "CA"
                        parts = address.strip().split()
                        for i, part in enumerate(parts):
                            if part == "CA" and i > 0:
                                city = parts[i - 1]
                                if i > 1 and parts[i - 2][0].isupper() and not parts[i - 2][0].isdigit():
                                    city = parts[i - 2] + " " + city
                                break

                event_type = "layoff"
                if layoff_type and "closure" in str(layoff_type).lower():
                    event_type = "facility_shutdown"

                results.append({
                    "company_name": company_name,
                    "event_type": event_type,
                    "event_date": event_date,
                    "employees_affected": emp_count,
                    "locations": [{"city": city, "state": "CA", "county": str(county or "").replace(" County", "")}] if county else [],
                    "source_url": CA_WARN_URL,
                    "raw_text": f"WARN notice: {company_name}, {emp_count} employees, {county or 'CA'}, {event_type}",
                })
            except Exception as e:
                logger.warning("warn_act.parse_error", error=str(e))
                continue

        wb.close()
        return results

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, max=16))
    async def _fetch_texas(self) -> list[dict]:
        """Fetch WARN notices from Texas via Socrata SODA API."""
        since = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%dT00:00:00")
        params = {
            "$where": f"notice_date > '{since}'",
            "$limit": "200",
            "$order": "notice_date DESC",
        }

        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.get(TX_WARN_URL, params=params)
            resp.raise_for_status()
            records = resp.json()

        results = []
        for record in records:
            company_name = record.get("job_site_name", "").strip()
            if not company_name:
                continue

            emp_count = 0
            emp_str = record.get("no_of_employees", "")
            if emp_str:
                try:
                    emp_count = int(str(emp_str).replace(",", "").strip())
                except (ValueError, TypeError):
                    emp_count = 0

            event_date = None
            date_str = record.get("layoff_date") or record.get("notice_date")
            if date_str:
                try:
                    event_date = datetime.strptime(date_str[:10], "%Y-%m-%d").date()
                except ValueError:
                    pass

            county = record.get("county", "").strip()
            type_code = str(record.get("type_code", "")).lower()
            event_type = "facility_shutdown" if "closure" in type_code else "layoff"

            results.append({
                "company_name": company_name,
                "event_type": event_type,
                "event_date": event_date,
                "employees_affected": emp_count,
                "locations": [{"city": "", "state": "TX", "county": county}] if county else [{"city": "", "state": "TX", "county": ""}],
                "source_url": TX_WARN_URL,
                "raw_text": f"WARN notice: {company_name}, {emp_count} employees, {county or 'TX'}, {event_type}",
            })

        return results

    @retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, max=16))
    async def _fetch_new_york(self) -> list[dict]:
        """Fetch WARN notices from New York DOL HTML page."""
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            resp = await client.get(NY_WARN_URL)
            resp.raise_for_status()
            html = resp.text

        results = []
        # NY DOL page contains a table with WARN notices
        # Extract table rows using regex (no BeautifulSoup dependency)
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.DOTALL | re.IGNORECASE)

        for row in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL | re.IGNORECASE)
            if len(cells) < 4:
                continue

            # Clean HTML tags from cell contents
            clean = lambda s: re.sub(r"<[^>]+>", "", s).strip()
            # Typical columns: Date, Company, Number Affected, Location/Region
            date_str = clean(cells[0])
            company_name = clean(cells[1])
            emp_str = clean(cells[2])
            location = clean(cells[3]) if len(cells) > 3 else ""

            if not company_name or company_name.lower() in ("company", "employer", ""):
                continue

            emp_count = 0
            if emp_str:
                # Extract first number from the cell
                match = re.search(r"(\d[\d,]*)", emp_str)
                if match:
                    try:
                        emp_count = int(match.group(1).replace(",", ""))
                    except ValueError:
                        emp_count = 0

            event_date = None
            if date_str:
                for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%B %d, %Y"):
                    try:
                        event_date = datetime.strptime(date_str.strip(), fmt).date()
                        break
                    except ValueError:
                        continue

            # Only include recent notices (within last 30 days)
            if event_date and (datetime.now().date() - event_date).days > 30:
                continue

            results.append({
                "company_name": company_name[:500],
                "event_type": "layoff",
                "event_date": event_date,
                "employees_affected": emp_count,
                "locations": [{"city": location, "state": "NY", "county": ""}],
                "source_url": NY_WARN_URL,
                "raw_text": f"WARN notice: {company_name}, {emp_count} employees, {location or 'NY'}, layoff",
            })

        return results
